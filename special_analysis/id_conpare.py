import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

class ExcelIDComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel ID 版本对比工具")
        self.root.geometry("700x500")
        self.root.resizable(False, False)
        
        self.old_file_path = None
        self.new_file_path = None
        
        self.setup_ui()
    
    def setup_ui(self):
        # 标题
        title_label = tk.Label(
            self.root, 
            text="Excel ID 版本对比分析工具", 
            font=("Arial", 16, "bold"),
            pady=20
        )
        title_label.pack()
        
        # 文件选择框架
        file_frame = ttk.LabelFrame(self.root, text="文件选择", padding=20)
        file_frame.pack(fill="x", padx=20, pady=10)
        
        # 旧版本文件
        old_frame = tk.Frame(file_frame)
        old_frame.pack(fill="x", pady=5)
        
        tk.Label(old_frame, text="旧版本文件:", width=12, anchor="w").pack(side="left")
        self.old_file_entry = tk.Entry(old_frame, width=40)
        self.old_file_entry.pack(side="left", padx=5)
        tk.Button(
            old_frame, 
            text="浏览", 
            command=lambda: self.select_file("old"),
            width=8
        ).pack(side="left")
        
        # 新版本文件
        new_frame = tk.Frame(file_frame)
        new_frame.pack(fill="x", pady=5)
        
        tk.Label(new_frame, text="新版本文件:", width=12, anchor="w").pack(side="left")
        self.new_file_entry = tk.Entry(new_frame, width=40)
        self.new_file_entry.pack(side="left", padx=5)
        tk.Button(
            new_frame, 
            text="浏览", 
            command=lambda: self.select_file("new"),
            width=8
        ).pack(side="left")
        
        # 分析按钮
        analyze_btn = tk.Button(
            self.root,
            text="开始分析",
            command=self.analyze_files,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 12, "bold"),
            width=20,
            height=2,
            cursor="hand2"
        )
        analyze_btn.pack(pady=20)
        
        # 结果显示框架
        result_frame = ttk.LabelFrame(self.root, text="分析结果", padding=10)
        result_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 结果文本框
        self.result_text = tk.Text(
            result_frame, 
            height=10, 
            width=80,
            font=("Courier", 10),
            wrap="word"
        )
        self.result_text.pack(side="left", fill="both", expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(result_frame, command=self.result_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.result_text.config(yscrollcommand=scrollbar.set)
    
    def select_file(self, file_type):
        """选择文件"""
        file_path = filedialog.askopenfilename(
            title=f"选择{'旧' if file_type == 'old' else '新'}版本Excel文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            if file_type == "old":
                self.old_file_path = file_path
                self.old_file_entry.delete(0, tk.END)
                self.old_file_entry.insert(0, file_path)
            else:
                self.new_file_path = file_path
                self.new_file_entry.delete(0, tk.END)
                self.new_file_entry.insert(0, file_path)
    
    def analyze_files(self):
        """分析文件差异"""
        # 清空结果显示
        self.result_text.delete(1.0, tk.END)
        
        # 验证文件是否选择
        if not self.old_file_path or not self.new_file_path:
            messagebox.showerror("错误", "请先选择两个Excel文件！")
            return
        
        try:
            # 读取Excel文件
            self.result_text.insert(tk.END, "正在读取文件...\n")
            self.root.update()
            
            old_df = pd.read_excel(self.old_file_path)
            new_df = pd.read_excel(self.new_file_path)
            
            # 获取B列数据（索引为1）
            old_ids = set(old_df.iloc[:, 1].dropna().astype(str))
            new_ids = set(new_df.iloc[:, 1].dropna().astype(str))
            
            # 计算差异
            deleted_ids = old_ids - new_ids  # 被删除的ID
            added_ids = new_ids - old_ids    # 新增的ID
            
            # 显示统计信息
            self.result_text.insert(tk.END, f"\n{'='*50}\n")
            self.result_text.insert(tk.END, f"旧版本ID总数: {len(old_ids)}\n")
            self.result_text.insert(tk.END, f"新版本ID总数: {len(new_ids)}\n")
            self.result_text.insert(tk.END, f"删除的ID数量: {len(deleted_ids)}\n")
            self.result_text.insert(tk.END, f"新增的ID数量: {len(added_ids)}\n")
            self.result_text.insert(tk.END, f"{'='*50}\n\n")
            
            # 生成结果文件
            self.result_text.insert(tk.END, "正在生成结果文件...\n")
            self.root.update()
            
            output_path = self.generate_result_file(
                old_df, new_df, deleted_ids, added_ids
            )
            
            self.result_text.insert(tk.END, f"\n✓ 分析完成！\n")
            self.result_text.insert(tk.END, f"结果已保存至: {output_path}\n")
            
            messagebox.showinfo("成功", f"分析完成！\n结果文件: {output_path}")
            
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时出错:\n{str(e)}")
            self.result_text.insert(tk.END, f"\n错误: {str(e)}\n")
    
    def generate_result_file(self, old_df, new_df, deleted_ids, added_ids):
        """生成结果Excel文件"""
        # 创建输出文件路径
        output_path = Path(self.old_file_path).parent / "ID对比结果.xlsx"
        
        # 创建Excel写入器
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 写入删除的ID（整行）
            deleted_rows = old_df[old_df.iloc[:, 1].astype(str).isin(deleted_ids)]
            deleted_rows.to_excel(writer, sheet_name='删除的ID', index=False)
            
            # 写入新增的ID（整行）
            added_rows = new_df[new_df.iloc[:, 1].astype(str).isin(added_ids)]
            added_rows.to_excel(writer, sheet_name='新增的ID', index=False)
            
            # 写入统计信息
            summary_data = {
                '项目': ['旧版本ID总数', '新版本ID总数', '删除的ID数量', '新增的ID数量'],
                '数量': [
                    len(set(old_df.iloc[:, 1].dropna().astype(str))),
                    len(set(new_df.iloc[:, 1].dropna().astype(str))),
                    len(deleted_ids),
                    len(added_ids)
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='统计信息', index=False)
        
        # 美化Excel
        self.beautify_excel(output_path)
        
        return output_path
    
    def beautify_excel(self, file_path):
        """美化Excel文件"""
        wb = openpyxl.load_workbook(file_path)
        
        # 定义样式
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        deleted_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        added_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        
        # 美化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # 根据工作表类型设置行颜色
            if sheet_name == '删除的ID':
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = deleted_fill
            elif sheet_name == '新增的ID':
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = added_fill
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)

def main():
    root = tk.Tk()
    app = ExcelIDComparator(root)
    root.mainloop()

if __name__ == "__main__":
    main()