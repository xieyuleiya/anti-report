"""
高速公路数据下载器（highway_downloader.py）

基于 optimized_download.py 创建，实现按高速公路、桥梁、测点的层级数据下载功能。

主要功能：
1. 按高速公路名称筛选数据（H列包含指定高速名）
2. 支持指定单个桥梁或下载高速下所有桥梁
3. 每个测点ID单独下载，按 H_F_I_B_D 格式命名文件
4. 异常处理：空数据文件标记为"异常"，记录详细统计信息
5. 文件存储结构：高速/桥名/测点

文件命名规则：
- 正常文件：H_F_I_B_D.txt
- 异常文件：H_F_I_B_D_异常.txt
- 其中：H=高速名, F=桥名, I=测点ID, B=墩号, D=厂家

存储结构：
ROAD/
└── 高速名/
    ├── 桥名1/
    │   ├── 测点ID1.txt
    │   ├── 测点ID2.txt
    │   └── 测点ID3_异常.txt
    └── 桥名2/
        ├── 测点ID4.txt
        └── 测点ID5.txt
"""

import json
import requests
import calendar
import zipfile
import os
import re
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import shutil
from collections import defaultdict

# 配置参数（可被 run_download 通过 apply_config 动态覆盖）
headers = {'Content-Type': 'application/json'}
API_URL = 'http://192.168.1.244:8122/InternalData/DataExport'
API_KEY = "337F11D3-7181-4570-9F44-CD42396BA266"  # 可由配置覆盖

# 必填配置（由主程序注入）；此处不再给出业务默认值，避免重复配置
EXCEL_PATH = None
ROAD = None
TARGET_HIGHWAYS = None
TARGET_HIGHWAY_BRIDGES = []  # [("高速", "桥梁"), ...]
USE_SPECIFIC_COMBINATIONS = False
START_DATE = None
END_DATE = None
BATCH_SIZE_DAYS = None

# ================= 配置注入 =================

def apply_config(cfg=None):
    """应用主程序注入的配置；若缺失关键项则抛出异常。"""
    global EXCEL_PATH, ROAD, TARGET_HIGHWAYS, TARGET_HIGHWAY_BRIDGES, USE_SPECIFIC_COMBINATIONS
    global START_DATE, END_DATE, BATCH_SIZE_DAYS
    if not cfg:
        raise ValueError("下载配置缺失：必须由主程序注入配置字典")
    # 必填项
    EXCEL_PATH = cfg['EXCEL_PATH']
    ROAD = cfg['ROAD']
    TARGET_HIGHWAYS = cfg['TARGET_HIGHWAYS']
    START_DATE = cfg['START_DATE']
    END_DATE = cfg['END_DATE']
    BATCH_SIZE_DAYS = cfg['BATCH_SIZE_DAYS']
    # 可选项
    TARGET_HIGHWAY_BRIDGES = cfg.get('TARGET_HIGHWAY_BRIDGES', [])
    USE_SPECIFIC_COMBINATIONS = len(TARGET_HIGHWAY_BRIDGES) > 0

# 统计信息收集
class DownloadStatistics:
    """下载统计信息收集类"""
    
    def __init__(self, highway_names):
        self.highway_names = highway_names
        self.highway_name = ", ".join(highway_names) if len(highway_names) > 1 else highway_names[0]
        self.total_bridges = 0
        self.total_sensors = 0
        self.download_success = 0
        self.download_failed = 0
        self.empty_files = 0
        self.heavy_missing_files = 0  # 大量缺失文件数
        self.light_missing_files = 0  # 少量缺失文件数
        self.bridge_details = {}  # {桥名: {测点数, 成功数, 失败数, 空文件数, 大量缺失数, 少量缺失数, 异常详情}}
        self.error_details = []   # 详细错误信息列表
        self.sensor_details = {}  # {桥名: {测点ID: {墩号, 厂家, 状态, 异常详情}}}
    
    def add_bridge(self, bridge_name, sensor_count):
        """添加桥梁信息"""
        self.total_bridges += 1
        self.total_sensors += sensor_count
        self.bridge_details[bridge_name] = {
            'sensor_count': sensor_count,
            'success_count': 0,
            'failed_count': 0,
            'empty_count': 0,
            'heavy_missing_count': 0,
            'light_missing_count': 0,
            'errors': []
        }
        self.sensor_details[bridge_name] = {}
    
    def record_download_result(self, bridge_name, sensor_id, success, is_abnormal=False, error_msg=None, abnormal_details=None, pier_number="", manufacturer="", sensor_number=""):
        """记录下载结果"""
        # 记录测点详细信息
        sensor_info = {
            'pier_number': pier_number,
            'manufacturer': manufacturer,
            'sensor_number': sensor_number,
            'status': 'success' if success and not is_abnormal else 'failed' if not success else 'abnormal',
            'error_msg': error_msg,
            'abnormal_details': abnormal_details
        }
        self.sensor_details[bridge_name][sensor_id] = sensor_info
        
        if success:
            if is_abnormal:
                # 根据异常类型进行分类统计
                if abnormal_details and abnormal_details.get('is_empty'):
                    # 空文件
                    self.empty_files += 1
                    self.bridge_details[bridge_name]['empty_count'] += 1
                elif abnormal_details and abnormal_details.get('time_check'):
                    # 时间完整性异常
                    tc = abnormal_details['time_check']
                    completeness_rate = tc.get('completeness_rate', 0)
                    if completeness_rate < 50:  # 大量缺失
                        self.heavy_missing_files += 1
                        self.bridge_details[bridge_name]['heavy_missing_count'] += 1
                    else:  # 少量缺失
                        self.light_missing_files += 1
                        self.bridge_details[bridge_name]['light_missing_count'] += 1
                else:
                    # 其他异常，计入空文件
                    self.empty_files += 1
                    self.bridge_details[bridge_name]['empty_count'] += 1
                
                # 记录异常详情
                if abnormal_details:
                    abnormal_info = f"{sensor_id}: {abnormal_details.get('reason', '未知异常')}"
                    if abnormal_details.get('time_check'):
                        tc = abnormal_details['time_check']
                        if tc.get('has_issues'):
                            abnormal_info += f" (完整性:{tc.get('completeness_rate', 0):.1f}%, 缺失:{tc.get('missing_count', 0)}天)"
                    self.bridge_details[bridge_name]['errors'].append(abnormal_info)
                    self.error_details.append(f"{bridge_name}-{abnormal_info}")
            else:
                self.download_success += 1
                self.bridge_details[bridge_name]['success_count'] += 1
        else:
            self.download_failed += 1
            self.bridge_details[bridge_name]['failed_count'] += 1
            if error_msg:
                self.bridge_details[bridge_name]['errors'].append(f"{sensor_id}: {error_msg}")
                self.error_details.append(f"{bridge_name}-{sensor_id}: {error_msg}")
    
    def generate_report(self):
        """生成统计报告"""
        report = []
        report.append(f"# {self.highway_name} 数据下载统计报告")
        report.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        
        # 总体统计
        report.append("## 总体统计")
        if len(self.highway_names) > 1:
            report.append(f"- 高速公路: {', '.join(self.highway_names)}")
        else:
            report.append(f"- 高速公路: {self.highway_name}")
        report.append(f"- 桥梁总数: {self.total_bridges}")
        report.append(f"- 测点总数: {self.total_sensors}")
        report.append(f"- 下载成功: {self.download_success}")
        report.append(f"- 下载失败: {self.download_failed}")
        report.append(f"- 空文件数: {self.empty_files}")
        report.append(f"- 大量缺失文件数: {self.heavy_missing_files} (完整性<50%)")
        report.append(f"- 少量缺失文件数: {self.light_missing_files} (完整性≥50%)")
        report.append(f"- 异常文件总数: {self.empty_files + self.heavy_missing_files + self.light_missing_files}")
        report.append("")
        
        # 时间完整性统计
        total_missing_files = self.heavy_missing_files + self.light_missing_files
        if total_missing_files > 0:
            report.append("## 时间完整性统计")
            report.append(f"- 大量缺失文件数: {self.heavy_missing_files} (完整性<50%)")
            report.append(f"- 少量缺失文件数: {self.light_missing_files} (完整性≥50%)")
            report.append(f"- 时间完整率: {((self.total_sensors - total_missing_files) / self.total_sensors * 100):.1f}%")
            report.append("")
        
        # 桥梁详细统计
        report.append("## 桥梁详细统计")
        report.append("| 桥梁名称 | 测点数 | 成功数 | 失败数 | 空文件 | 大量缺失 | 少量缺失 | 状态 |")
        report.append("|:---------|:------:|:------:|:------:|:------:|:--------:|:--------:|:----:|")
        
        for bridge_name, details in self.bridge_details.items():
            total_abnormal = details['empty_count'] + details['heavy_missing_count'] + details['light_missing_count']
            status = "✅ 正常" if details['failed_count'] == 0 and total_abnormal == 0 else "⚠️ 部分异常" if details['success_count'] > 0 else "❌ 全部失败"
            report.append(f"| {bridge_name} | {details['sensor_count']} | {details['success_count']} | {details['failed_count']} | {details['empty_count']} | {details['heavy_missing_count']} | {details['light_missing_count']} | {status} |")
        
        # 测点详细统计
        report.append("")
        report.append("## 测点详细统计")
        for bridge_name, sensors in self.sensor_details.items():
            report.append(f"### {bridge_name}")
            report.append("| 测点ID | 测点编号 | 墩号 | 厂家 | 状态 | 详情 |")
            report.append("|:-------|:---------|:-----|:-----|:----:|:----|")
            
            for sensor_id, info in sensors.items():
                status_emoji = {
                    'success': '✅',
                    'failed': '❌', 
                    'abnormal': '⚠️'
                }.get(info['status'], '❓')
                
                status_text = {
                    'success': '成功',
                    'failed': '失败',
                    'abnormal': '异常'
                }.get(info['status'], '未知')
                
                details = ""
                if info['status'] == 'failed' and info['error_msg']:
                    details = f"下载失败: {info['error_msg']}"
                elif info['status'] == 'abnormal' and info['abnormal_details']:
                    abnormal = info['abnormal_details']
                    if abnormal.get('time_check'):
                        tc = abnormal['time_check']
                        if tc.get('has_issues'):
                            details = f"时间不完整: {tc.get('completeness_rate', 0):.1f}%, 缺失{tc.get('missing_count', 0)}天"
                            if tc.get('missing_dates'):
                                missing_dates = [str(d) for d in tc['missing_dates'][:3]]  # 只显示前3个
                                if len(tc['missing_dates']) > 3:
                                    missing_dates.append(f"...(共{tc['missing_count']}天)")
                                details += f" [{', '.join(missing_dates)}]"
                        else:
                            details = abnormal.get('reason', '未知异常')
                    else:
                        details = abnormal.get('reason', '未知异常')
                
                report.append(f"| {sensor_id} | {info['sensor_number']} | {info['pier_number']} | {info['manufacturer']} | {status_emoji} {status_text} | {details} |")
            report.append("")
        
        # 异常详情汇总
        if self.error_details:
            report.append("## 异常详情汇总")
            for i, error in enumerate(self.error_details, 1):
                report.append(f"{i}. {error}")
        
        return "\n".join(report)
    
    def save_report(self, save_path):
        """保存报告到文件"""
        # 保存Markdown报告
        report_content = self.generate_report()
        report_file = os.path.join(save_path, f"{self.highway_name}_下载报告.md")
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        # 保存Excel报告
        excel_file = self.save_excel_report(save_path)
        
        print(f"📄 统计报告已保存: {report_file}")
        print(f"📊 Excel报告已保存: {excel_file}")
        return report_file
    
    def save_excel_report(self, save_path):
        """保存Excel格式的详细报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # 总体统计工作表
        ws_summary = wb.active
        ws_summary.title = "总体统计"
        
        # 设置标题
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        ws_summary['A1'] = f"{self.highway_name} 数据下载统计报告"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 总体统计
        ws_summary['A4'] = "总体统计"
        ws_summary['A4'].font = header_font
        
        summary_data = [
            ["高速公路", self.highway_name],
            ["桥梁总数", self.total_bridges],
            ["测点总数", self.total_sensors],
            ["下载成功", self.download_success],
            ["下载失败", self.download_failed],
            ["空文件数", self.empty_files],
            ["大量缺失文件数", self.heavy_missing_files],
            ["少量缺失文件数", self.light_missing_files],
            ["异常文件总数", self.empty_files + self.heavy_missing_files + self.light_missing_files],
        ]
        
        for i, (label, value) in enumerate(summary_data, 5):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
        
        # 桥梁详细统计工作表
        ws_bridges = wb.create_sheet("桥梁详细统计")
        
        # 设置表头
        headers = ["桥梁名称", "测点数", "成功数", "失败数", "空文件", "大量缺失", "少量缺失", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws_bridges.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 填充数据
        row = 2
        for bridge_name, details in self.bridge_details.items():
            total_abnormal = details['empty_count'] + details['heavy_missing_count'] + details['light_missing_count']
            status = "✅ 正常" if details['failed_count'] == 0 and total_abnormal == 0 else "⚠️ 部分异常" if details['success_count'] > 0 else "❌ 全部失败"
            
            ws_bridges.cell(row=row, column=1, value=bridge_name)
            ws_bridges.cell(row=row, column=2, value=details['sensor_count'])
            ws_bridges.cell(row=row, column=3, value=details['success_count'])
            ws_bridges.cell(row=row, column=4, value=details['failed_count'])
            ws_bridges.cell(row=row, column=5, value=details['empty_count'])
            ws_bridges.cell(row=row, column=6, value=details['heavy_missing_count'])
            ws_bridges.cell(row=row, column=7, value=details['light_missing_count'])
            ws_bridges.cell(row=row, column=8, value=status)
            row += 1
        
        # 测点详细统计工作表
        ws_sensors = wb.create_sheet("测点详细统计")
        
        # 设置表头
        sensor_headers = ["桥梁名称", "测点ID", "测点编号", "墩号", "厂家", "状态", "详情"]
        for col, header in enumerate(sensor_headers, 1):
            cell = ws_sensors.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 填充数据
        row = 2
        for bridge_name, sensors in self.sensor_details.items():
            for sensor_id, info in sensors.items():
                status_text = {
                    'success': '✅ 成功',
                    'failed': '❌ 失败',
                    'abnormal': '⚠️ 异常'
                }.get(info['status'], '❓ 未知')
                
                details = ""
                if info['status'] == 'failed' and info['error_msg']:
                    details = f"下载失败: {info['error_msg']}"
                elif info['status'] == 'abnormal' and info['abnormal_details']:
                    abnormal = info['abnormal_details']
                    if abnormal.get('time_check'):
                        tc = abnormal['time_check']
                        if tc.get('has_issues'):
                            details = f"时间不完整: {tc.get('completeness_rate', 0):.1f}%, 缺失{tc.get('missing_count', 0)}天"
                            if tc.get('missing_dates'):
                                missing_dates = [str(d) for d in tc['missing_dates'][:3]]
                                if len(tc['missing_dates']) > 3:
                                    missing_dates.append(f"...(共{tc['missing_count']}天)")
                                details += f" [{', '.join(missing_dates)}]"
                        else:
                            details = abnormal.get('reason', '未知异常')
                    else:
                        details = abnormal.get('reason', '未知异常')
                
                ws_sensors.cell(row=row, column=1, value=bridge_name)
                ws_sensors.cell(row=row, column=2, value=sensor_id)
                ws_sensors.cell(row=row, column=3, value=info['sensor_number'])
                ws_sensors.cell(row=row, column=4, value=info['pier_number'])
                ws_sensors.cell(row=row, column=5, value=info['manufacturer'])
                ws_sensors.cell(row=row, column=6, value=status_text)
                ws_sensors.cell(row=row, column=7, value=details)
                row += 1
        
        # 异常详情工作表
        if self.error_details:
            ws_errors = wb.create_sheet("异常详情")
            
            # 设置表头
            error_headers = ["序号", "异常信息"]
            for col, header in enumerate(error_headers, 1):
                cell = ws_errors.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 填充数据
            for i, error in enumerate(self.error_details, 2):
                ws_errors.cell(row=i, column=1, value=i-1)
                ws_errors.cell(row=i, column=2, value=error)
        
        # 调整列宽
        for ws in [ws_summary, ws_bridges, ws_sensors, ws_errors]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        excel_file = os.path.join(save_path, f"{self.highway_name}_下载报告.xlsx")
        wb.save(excel_file)
        return excel_file


def create_directory(path):
    """创建目录"""
    if not os.path.exists(path):
        os.makedirs(path)
        print(f"📁 创建目录: {path}")


def clean_directory(directory):
    """清理目录中的旧文件"""
    if os.path.exists(directory):
        for file_name in os.listdir(directory):
            file_path = os.path.join(directory, file_name)
            if os.path.isfile(file_path):
                os.remove(file_path)
        print(f"🧹 清理目录: {directory}")


def download_sensor_data(sensor_id, start_date, end_date, save_dir, batch_suffix="", pier_number="", sensor_number=""):
    """
    下载单个测点的数据
    
    参数:
    - sensor_id: 测点ID
    - start_date: 开始日期
    - end_date: 结束日期
    - save_dir: 保存目录
    - batch_suffix: 批次后缀
    - pier_number: 墩号 (用于显示)
    - sensor_number: 测点编号 (用于显示)
    
    返回:
    - tuple: (是否成功, 是否为空文件, 错误信息)
    """
    data = {
        "ids": str(sensor_id),
        "start": f"{start_date} 00:00:00",
        "end": f"{end_date} 23:59:59",
        "key": API_KEY,
        "type": 1
    }
    
    try:
        pier_info = f" (墩号:{pier_number})" if pier_number else ""
        sensor_info = f" (测点编号:{sensor_number})" if sensor_number and sensor_number != "未知" else ""
        print(f"   🔄 下载测点 {sensor_id}{sensor_info}{pier_info}: {start_date} 到 {end_date}")
        
        # 使用与api_test.py相同的超时设置
        response = requests.post(API_URL, data=json.dumps(data), headers=headers, timeout=10)
        
        if response.status_code == 200:
            # 检查响应内容大小
            if len(response.content) < 100:
                return True, True, "数据为空或过小"
            
            # 生成文件名
            filename = f"{sensor_id}{batch_suffix}.zip"
            save_path = os.path.join(save_dir, filename)
            
            with open(save_path, 'wb') as file:
                file.write(response.content)
            
            file_size_mb = len(response.content) / (1024 * 1024)
            print(f"   ✅ 下载成功: {filename} ({file_size_mb:.2f} MB)")
            
            return True, False, None
        else:
            error_msg = f"HTTP {response.status_code}"
            print(f"   ❌ 下载失败: {error_msg}")
            return False, False, error_msg
            
    except Exception as e:
        error_msg = str(e)
        print(f"   ❌ 下载异常: {error_msg}")
        return False, False, error_msg


def download_sensor_batch(sensor_id, start_date, end_date, save_dir, pier_number="", sensor_number=""):
    """
    分批下载测点数据
    
    参数:
    - sensor_id: 测点ID
    - start_date: 开始日期
    - end_date: 结束日期
    - save_dir: 保存目录
    - pier_number: 墩号 (用于显示)
    - sensor_number: 测点编号 (用于显示)
    
    返回:
    - tuple: (是否成功, 是否为空文件, 错误信息)
    """
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    total_days = (end_dt - start_dt).days + 1
    
    if total_days <= BATCH_SIZE_DAYS:
        # 小时间范围，直接下载
        return download_sensor_data(sensor_id, start_date, end_date, save_dir, "", pier_number, sensor_number)
    else:
        # 大时间范围，分批下载
        print(f"   ⚠️  时间范围较大({total_days}天)，分批下载")
        current_start = start_dt
        batch_count = 0
        all_empty = True
        any_success = False
        
        while current_start <= end_dt:
            current_end = min(current_start + timedelta(days=BATCH_SIZE_DAYS - 1), end_dt)
            batch_count += 1
            
            batch_suffix = f"_batch{batch_count:02d}"
            success, is_empty, error = download_sensor_data(
                sensor_id,
                current_start.strftime("%Y-%m-%d"),
                current_end.strftime("%Y-%m-%d"),
                save_dir,
                batch_suffix,
                pier_number,
                sensor_number
            )
            
            if success:
                any_success = True
                if not is_empty:
                    all_empty = False
            
            # 移动到下一批次
            current_start = current_end + timedelta(days=1)
            
            # 增加请求间隔，避免触发服务器限制
            if current_start <= end_dt:
                import time
                time.sleep(2.0)  # 增加到2秒
        
        if any_success:
            return True, all_empty, None
        else:
            return False, False, "所有批次都失败"


def check_time_completeness(data_lines, start_date, end_date):
    """
    检查时间完整性
    
    参数:
    - data_lines: 数据行列表
    - start_date: 期望的开始日期 (YYYY-MM-DD)
    - end_date: 期望的结束日期 (YYYY-MM-DD)
    
    返回:
    - dict: 检查结果详情
    """
    try:
        # 解析期望的日期范围
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        expected_days = (end_dt - start_dt).days + 1
        
        # 提取数据中的日期
        data_dates = set()
        valid_lines = 0
        
        for line in data_lines[1:]:  # 跳过表头
            line = line.strip()
            if not line:
                continue
                
            try:
                # 尝试解析时间戳（假设前两列是日期和时间）
                parts = line.split()
                if len(parts) >= 2:
                    date_str = parts[0]
                    time_str = parts[1]
                    datetime_str = f"{date_str} {time_str}"
                    
                    # 尝试多种时间格式
                    parsed_dt = None
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y/%m/%d %H:%M:%S']:
                        try:
                            parsed_dt = datetime.strptime(datetime_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if parsed_dt:
                        data_dates.add(parsed_dt.date())
                        valid_lines += 1
            except Exception:
                continue
        
        # 统计缺失的日期
        expected_dates = set()
        current_dt = start_dt
        while current_dt <= end_dt:
            expected_dates.add(current_dt.date())
            current_dt += timedelta(days=1)
        
        missing_dates = expected_dates - data_dates
        extra_dates = data_dates - expected_dates
        
        # 计算完整性指标
        actual_days = len(data_dates)
        missing_count = len(missing_dates)
        extra_count = len(extra_dates)
        completeness_rate = (actual_days / expected_days) * 100 if expected_days > 0 else 0
        
        # 判断是否有问题
        has_issues = missing_count > 0 or extra_count > 0 or actual_days == 0
        
        # 构建详细报告
        summary_parts = []
        if actual_days == 0:
            summary_parts.append("无有效数据")
        else:
            summary_parts.append(f"数据完整性: {completeness_rate:.1f}%")
            if missing_count > 0:
                summary_parts.append(f"缺失{missing_count}天")
            if extra_count > 0:
                summary_parts.append(f"多余{extra_count}天")
        
        summary = ", ".join(summary_parts)
        
        return {
            'has_issues': has_issues,
            'expected_days': expected_days,
            'actual_days': actual_days,
            'missing_count': missing_count,
            'extra_count': extra_count,
            'completeness_rate': completeness_rate,
            'missing_dates': sorted(list(missing_dates)),
            'extra_dates': sorted(list(extra_dates)),
            'data_dates': sorted(list(data_dates)),
            'valid_lines': valid_lines,
            'summary': summary
        }
        
    except Exception as e:
        return {
            'has_issues': True,
            'error': f"时间检查失败: {str(e)}",
            'summary': f"时间检查异常: {str(e)}"
        }


def process_sensor_files(sensor_dir, sensor_id, bridge_name, highway_name, pier_number, manufacturer, sensor_number, start_date=None, end_date=None):
    """
    处理测点下载的文件，解压并重命名，并进行时间完整性检查
    
    参数:
    - sensor_dir: 测点目录
    - sensor_id: 测点ID
    - bridge_name: 桥梁名称
    - highway_name: 高速公路名称
    - pier_number: 墩号
    - manufacturer: 厂家
    - sensor_number: 测点编号（I列）
    - start_date: 下载开始日期 (用于时间完整性检查)
    - end_date: 下载结束日期 (用于时间完整性检查)
    
    返回:
    - tuple: (是否成功, 是否为空文件, 错误信息, 异常详情)
    """
    try:
        # 查找ZIP文件
        zip_files = [f for f in os.listdir(sensor_dir) if f.endswith('.zip')]
        if not zip_files:
            return False, False, "未找到ZIP文件", None
        
        # 解压文件
        extracted_files = []
        for zip_file in zip_files:
            zip_path = os.path.join(sensor_dir, zip_file)
            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    for fn in zf.namelist():
                        if fn.endswith('.txt'):
                            extracted_path = zf.extract(fn, sensor_dir)
                            extracted_files.append(extracted_path)
                # 删除ZIP文件
                os.remove(zip_path)
            except Exception as e:
                return False, False, f"解压失败: {str(e)}", None
        
        if not extracted_files:
            return True, True, "解压后无有效文件", None
        
        # 合并所有文件内容
        all_data = []
        header_added = False
        total_lines = 0
        
        for file_path in extracted_files:
            try:
                # 尝试多种编码方式读取文件
                lines = None
                for encoding in ['utf-8-sig', 'utf-8', 'gb18030', 'gbk']:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                        break
                    except UnicodeDecodeError:
                        continue
                
                if lines:
                    # 清理第一行，去除可能的BOM标记和特殊字符
                    if not header_added:
                        first_line = lines[0].strip()
                        # 去除开头的特殊字符，只保留时间戳开始的部分
                        if first_line:
                            # 查找第一个时间戳格式的位置
                            import re
                            timestamp_match = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', first_line)
                            if timestamp_match:
                                start_pos = timestamp_match.start()
                                cleaned_line = first_line[start_pos:] + '\n'
                                all_data.append(cleaned_line)
                            else:
                                all_data.append(first_line + '\n')
                        header_added = True
                    
                    # 处理后续行
                    for line in lines[1:]:
                        line = line.strip()
                        if line:  # 只添加非空行
                            all_data.append(line + '\n')
                            total_lines += 1
            except Exception:
                continue
            finally:
                # 删除临时文件
                try:
                    os.remove(file_path)
                except:
                    pass
        
        # 判断是否为空文件
        is_empty = total_lines <= 0  # 只有标题行或完全为空
        
        # 时间完整性检查
        time_check_result = None
        if not is_empty and start_date and end_date:
            time_check_result = check_time_completeness(all_data, start_date, end_date)
        
        # 根据检查结果决定文件名和异常标记
        is_abnormal = is_empty or (time_check_result and time_check_result['has_issues'])
        
        if is_abnormal:
            # 异常文件，根据缺失程度进行细分标记
            if is_empty:
                final_filename = f"{sensor_id}_{sensor_number}_{highway_name}_{bridge_name}_{pier_number}_{manufacturer}_空文件.txt"
                abnormal_reason = "文件为空"
            elif time_check_result and time_check_result['has_issues']:
                completeness_rate = time_check_result.get('completeness_rate', 0)
                missing_count = time_check_result.get('missing_count', 0)
                
                if completeness_rate < 50:  # 缺失超过50%
                    final_filename = f"{sensor_id}_{sensor_number}_{highway_name}_{bridge_name}_{pier_number}_{manufacturer}_大量缺失.txt"
                    abnormal_reason = f"大量缺失: {completeness_rate:.1f}%, 缺失{missing_count}天"
                else:  # 缺失少于50%
                    final_filename = f"{sensor_id}_{sensor_number}_{highway_name}_{bridge_name}_{pier_number}_{manufacturer}_少量缺失.txt"
                    abnormal_reason = f"少量缺失: {completeness_rate:.1f}%, 缺失{missing_count}天"
            else:
                final_filename = f"{sensor_id}_{sensor_number}_{highway_name}_{bridge_name}_{pier_number}_{manufacturer}_异常.txt"
                abnormal_reason = "未知异常"
        else:
            # 正常文件
            final_filename = f"{sensor_id}_{sensor_number}_{highway_name}_{bridge_name}_{pier_number}_{manufacturer}.txt"
            abnormal_reason = None
        
        final_path = os.path.join(sensor_dir, final_filename)
        
        # 保存合并后的文件
        # 即使数据为空也要创建文件，保持文件结构完整性
        with open(final_path, 'w', encoding='utf-8', newline='') as f:
            if all_data:
                f.writelines(all_data)
            # 如果数据为空，文件仍然会被创建，只是内容为空
        
        # 构建异常详情
        abnormal_details = None
        if is_abnormal:
            abnormal_details = {
                'is_empty': is_empty,
                'total_lines': total_lines,
                'time_check': time_check_result,
                'reason': abnormal_reason
            }
        
        return True, is_empty, abnormal_reason, abnormal_details
        
    except Exception as e:
        return False, False, f"处理文件失败: {str(e)}", None


def main():
    """主函数"""
    # 从命令行参数读取配置
    config = {
        'EXCEL_PATH': EXCEL_PATH,
        'ROAD': ROAD,
        'TARGET_HIGHWAYS': TARGET_HIGHWAYS,
        'TARGET_HIGHWAY_BRIDGES': TARGET_HIGHWAY_BRIDGES,
        'START_DATE': START_DATE,
        'END_DATE': END_DATE,
        'BATCH_SIZE_DAYS': BATCH_SIZE_DAYS,
    }
    apply_config(config)

    print(f"🚀 开始下载数据...")
    print(f"📋 配置信息:")
    print(f"  - Excel文件: {EXCEL_PATH}")
    print(f"  - 保存路径: {ROAD}")
    print(f"  - 目标高速公路: {', '.join(TARGET_HIGHWAYS)}")
    if USE_SPECIFIC_COMBINATIONS:
        print(f"  - 目标组合: {len(TARGET_HIGHWAY_BRIDGES)} 个")
        for i, (highway, bridge) in enumerate(TARGET_HIGHWAY_BRIDGES, 1):
            print(f"    {i}. {highway} - {bridge}")
    else:
        print(f"  - 目标高速公路: {', '.join(TARGET_HIGHWAYS)}")
    
    print(f"  - 下载时间范围: {START_DATE} 到 {END_DATE}")
    print(f"  - 批量大小: {BATCH_SIZE_DAYS} 天")
    
    # 初始化统计信息
    if USE_SPECIFIC_COMBINATIONS:
        highway_names = list(set([h for h, b in TARGET_HIGHWAY_BRIDGES]))
        stats = DownloadStatistics(highway_names)
    else:
        stats = DownloadStatistics(TARGET_HIGHWAYS)
    
    # 读取Excel文件
    try:
        df = pd.read_excel(EXCEL_PATH)
        print(f"✅ 成功读取Excel文件，共 {len(df)} 行数据")
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {str(e)}")
        return
    
    # 根据配置方式筛选数据
    if USE_SPECIFIC_COMBINATIONS:
        # 方式2: 筛选指定的高速-桥梁组合
        combined_mask = pd.Series([False] * len(df))
        for target_highway, target_bridge in TARGET_HIGHWAY_BRIDGES:
            mask = (df.iloc[:, 6].astype(str).str.contains(target_highway, na=False) & 
                   df.iloc[:, 5].astype(str).str.contains(target_bridge, na=False))
            combined_mask = combined_mask | mask
        
        highway_data = df[combined_mask]
        print(f"📊 找到 {len(highway_data)} 条匹配指定组合的数据")
        
        if highway_data.empty:
            print(f"❌ 未找到匹配指定组合的数据")
            return
    else:
        # 方式1: 筛选指定高速公路的数据（G列包含任一高速名）
        highway_mask = df.iloc[:, 6].astype(str).str.contains('|'.join(TARGET_HIGHWAYS), na=False)
        highway_data = df[highway_mask]
        print(f"📊 找到 {len(highway_data)} 条包含关键词 '{', '.join(TARGET_HIGHWAYS)}' 的数据")
        
        if highway_data.empty:
            print(f"❌ 未找到包含关键词 '{', '.join(TARGET_HIGHWAYS)}' 的数据")
            return
    
    # 使用筛选后的数据
    bridge_data = highway_data
    print(f"📊 下载所有桥梁数据，共 {len(bridge_data)} 条")
    
    # 按高速公路和桥梁分组
    bridge_groups = bridge_data.groupby([bridge_data.iloc[:, 6], bridge_data.iloc[:, 5]])  # G列(高速名), F列(桥名)
    print(f"📊 找到 {len(bridge_groups)} 个高速-桥梁组合")
    
    # 为每个高速-桥梁组合创建目录并下载数据
    for (excel_highway_name, bridge_name), bridge_group in bridge_groups:
        if pd.isna(bridge_name) or bridge_name == '':
            continue
        
        # 确定使用哪个高速名来创建文件夹
        folder_highway_name = None
        for target_highway in TARGET_HIGHWAYS:
            if target_highway in excel_highway_name:
                folder_highway_name = target_highway
                break
        
        if folder_highway_name is None:
            print(f"   ⚠️  跳过未匹配的高速: {excel_highway_name}")
            continue
        
        print(f"\n🔧 处理高速-桥梁: {folder_highway_name} - {bridge_name}")
        
        # 统计该桥梁的测点数
        sensor_count = len(bridge_group)
        stats.add_bridge(f"{folder_highway_name}-{bridge_name}", sensor_count)
        
        # 创建桥梁目录 - 使用指定的高速名
        bridge_dir = os.path.join(ROAD, folder_highway_name, bridge_name)
        create_directory(bridge_dir)
        
        # 清理旧文件
        clean_directory(bridge_dir)
        
        # 处理每个测点
        for idx, row in bridge_group.iterrows():
            sensor_id_raw = row.iloc[7]  # H列是测点ID
            sensor_number_raw = row.iloc[8]  # I列是测点编号
            
            # 清理测点ID格式，去除小数点，转换为干净字符串
            if pd.isna(sensor_id_raw):
                print(f"   ⚠️  跳过无效测点ID: {sensor_id_raw}")
                continue
            
            # 转换为字符串并清理格式
            sensor_id = str(int(float(sensor_id_raw))) if not pd.isna(sensor_id_raw) else ""
            if not sensor_id or sensor_id == 'nan':
                print(f"   ⚠️  跳过无效测点ID: {sensor_id_raw}")
                continue
            
            # 获取测点编号（I列）
            sensor_number = str(sensor_number_raw) if not pd.isna(sensor_number_raw) else "未知"
            if sensor_number == 'nan':
                sensor_number = "未知"
            
            pier_number = str(row.iloc[1]) if len(row) > 1 else "未知"  # B列是墩号
            manufacturer = str(row.iloc[3]) if len(row) > 3 else "未知"  # D列是厂家
            
            print(f"   📡 处理测点 {sensor_id} (测点编号:{sensor_number}, 墩号:{pier_number}) ({idx + 1}/{sensor_count}) - 厂家:{manufacturer}")
            
            # 创建测点目录
            sensor_dir = os.path.join(bridge_dir, sensor_id)
            create_directory(sensor_dir)
            
            # 下载数据
            success, is_empty, error = download_sensor_batch(sensor_id, START_DATE, END_DATE, sensor_dir, pier_number, sensor_number)
            
            if success:
                # 处理下载的文件 - 使用指定的高速名
                process_success, process_empty, process_error, abnormal_details = process_sensor_files(
                    sensor_dir, sensor_id, bridge_name, folder_highway_name, pier_number, manufacturer, sensor_number,
                    START_DATE, END_DATE
                )
                
                if process_success:
                    # 记录结果
                    is_abnormal = is_empty or process_empty or (abnormal_details and abnormal_details.get('reason'))
                    stats.record_download_result(f"{folder_highway_name}-{bridge_name}", sensor_id, True, 
                                               is_abnormal, 
                                               process_error,
                                               abnormal_details,
                                               pier_number,
                                               manufacturer,
                                               sensor_number)
                    
                    # 如果有异常详情，记录详细信息
                    if abnormal_details:
                        print(f"   📊 时间完整性检查: {abnormal_details.get('reason', '正常')}")
                        if abnormal_details.get('time_check'):
                            tc = abnormal_details['time_check']
                            if tc.get('has_issues'):
                                print(f"      - 期望天数: {tc.get('expected_days', 'N/A')}")
                                print(f"      - 实际天数: {tc.get('actual_days', 'N/A')}")
                                print(f"      - 缺失天数: {tc.get('missing_count', 'N/A')}")
                                print(f"      - 多余天数: {tc.get('extra_count', 'N/A')}")
                                if tc.get('missing_dates'):
                                    missing_dates_str = ', '.join([str(d) for d in tc['missing_dates'][:5]])  # 只显示前5个
                                    if len(tc['missing_dates']) > 5:
                                        missing_dates_str += f" ... (共{tc['missing_count']}天)"
                                    print(f"      - 缺失日期: {missing_dates_str}")
                    
                    # 将处理后的文件移动到桥梁目录，而不是删除
                    try:
                        # 查找处理后的文件
                        for file_name in os.listdir(sensor_dir):
                            if file_name.endswith('.txt'):
                                source_path = os.path.join(sensor_dir, file_name)
                                dest_path = os.path.join(bridge_dir, file_name)
                                # 移动文件到桥梁目录
                                shutil.move(source_path, dest_path)
                                print(f"   📁 文件已保存: {file_name}")
                    except Exception as e:
                        print(f"   ⚠️  移动文件失败: {e}")
                    finally:
                        # 清理临时测点目录
                        try:
                            shutil.rmtree(sensor_dir)
                        except:
                            pass
                else:
                    # 处理失败
                    stats.record_download_result(f"{folder_highway_name}-{bridge_name}", sensor_id, False, False, process_error, None, pier_number, manufacturer, sensor_number)
            else:
                # 下载失败
                stats.record_download_result(f"{folder_highway_name}-{bridge_name}", sensor_id, False, False, error, None, pier_number, manufacturer, sensor_number)
            
            # 在测点之间增加延迟，避免请求过于频繁
            if idx < len(bridge_group) - 1:  # 不是最后一个测点
                import time
                time.sleep(1.0)  # 每个测点间隔1秒
    
    # 生成统计报告
    print(f"\n📊 下载完成，生成统计报告...")
    report_path = stats.save_report(ROAD)
    
    print(f"\n🎉 {', '.join(TARGET_HIGHWAYS)} 数据下载完成！")
    print(f"📄 详细报告: {report_path}")


def run_download(config=None):
    """供外部调用的下载入口，支持传入配置覆盖默认值。"""
    apply_config(config)
    # 直接调用 main() 使用当前模块级变量执行
    return main()


if __name__ == "__main__":
    main() 