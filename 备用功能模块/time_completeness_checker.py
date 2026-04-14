# -*- coding: utf-8 -*-
"""
时间完整性检查模块
检查文件夹中所有文件的时间完整性，生成详细的检查报告
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TimeCompletenessChecker:
    """时间完整性检查器（按天检查）"""
    
    def __init__(self, folder_path: str, start_date: str, end_date: str):
        """
        初始化检查器
        
        Args:
            folder_path: 要检查的文件夹路径
            start_date: 开始日期，格式：YYYY-MM-DD
            end_date: 结束日期，格式：YYYY-MM-DD
        """
        self.folder_path = Path(folder_path)
        self.start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        self.end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        # 验证文件夹路径
        if not self.folder_path.exists():
            raise ValueError(f"文件夹路径不存在: {folder_path}")
        if not self.folder_path.is_dir():
            raise ValueError(f"路径不是文件夹: {folder_path}")
    
    def check_all_files(self) -> Dict:
        """
        检查文件夹中所有文件的时间完整性（按天检查）
        
        Returns:
            包含所有检查结果的字典
        """
        print(f"🔍 开始检查文件夹: {self.folder_path}")
        print(f"📅 时间范围: {self.start_date} 到 {self.end_date}")
        print("=" * 80)
        
        # 查找所有文本文件
        txt_files = list(self.folder_path.glob("*.txt"))
        
        if not txt_files:
            print("⚠️ 未找到任何 .txt 文件")
            return {}
        
        print(f"📁 找到 {len(txt_files)} 个文件")
        print("-" * 80)
        
        all_results = {}
        
        for file_path in txt_files:
            print(f"📊 正在检查: {file_path.name}")
            try:
                result = self._check_single_file(file_path)
                all_results[file_path.name] = result
                print(f"   ✅ 完成: {result['total_records']} 条记录, "
                      f"缺失 {result['missing_count']} 天")
            except Exception as e:
                print(f"   ❌ 检查失败: {str(e)}")
                all_results[file_path.name] = {
                    'error': str(e),
                    'total_records': 0,
                    'missing_count': 0
                }
            print()
        
        return all_results
    
    def _check_single_file(self, file_path: Path) -> Dict:
        """
        检查单个文件的时间完整性（按天检查）
        
        Args:
            file_path: 文件路径
            
        Returns:
            检查结果字典
        """
        # 读取文件
        try:
            # 尝试不同的分隔符和编码
            data = self._read_data_file(file_path)
        except Exception as e:
            raise Exception(f"读取文件失败: {str(e)}")
        
        if data.empty:
            return {
                'total_records': 0,
                'missing_count': 0,
                'missing_dates': [],
                'daily_stats': {},
                'data_stats': {},
                'error': '文件为空'
            }
        
        # 解析时间列，提取日期
        dates = self._parse_dates(data)
        
        if len(dates) == 0:
            return {
                'total_records': 0,
                'missing_count': 0,
                'missing_dates': [],
                'daily_stats': {},
                'data_stats': {},
                'error': '无法解析时间戳'
            }
        
        # 按天统计数据量
        daily_stats = self._calculate_daily_stats(dates)
        
        # 检查缺失的日期
        missing_dates = self._find_missing_dates(daily_stats)
        
        # 数据统计信息
        data_stats = self._calculate_data_stats(data)
        
        return {
            'total_records': len(dates),
            'missing_count': len(missing_dates),
            'missing_dates': missing_dates,
            'daily_stats': daily_stats,
            'data_stats': data_stats,
            'file_path': str(file_path)
        }
    
    def _read_data_file(self, file_path: Path) -> pd.DataFrame:
        """
        读取数据文件，尝试多种格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            DataFrame
        """
        # 尝试不同的编码和分隔符
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
        separators = ['\t', ' ', ',', '\s+']
        
        for encoding in encodings:
            for sep in separators:
                try:
                    if sep == '\s+':
                        data = pd.read_csv(file_path, sep=sep, header=None, 
                                          encoding=encoding, engine='python',
                                          on_bad_lines='skip')
                    else:
                        data = pd.read_csv(file_path, sep=sep, header=None, 
                                          encoding=encoding, on_bad_lines='skip')
                    
                    if not data.empty and len(data.columns) >= 2:
                        return data
                except:
                    continue
        
        # 如果都失败了，尝试按行读取
        try:
            lines = []
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        lines.append(line.split())
            
            if lines:
                max_cols = max(len(line) for line in lines)
                data = pd.DataFrame(lines)
                return data
        except:
            pass
        
        return pd.DataFrame()
    
    def _parse_dates(self, data: pd.DataFrame) -> List[datetime]:
        """
        解析时间戳并提取日期
        
        Args:
            data: DataFrame
            
        Returns:
            日期时间列表（用于统计）
        """
        timestamps = []
        
        # 尝试从第一列解析时间
        time_column = data.iloc[:, 0]
        
        # 尝试多种时间格式
        time_formats = [
            '%Y-%m-%d %H:%M:%S.%f',  # 2025-01-01 00:14:00.000
            '%Y-%m-%d %H:%M:%S',      # 2025-01-01 00:14:00
            '%Y/%m/%d %H:%M:%S.%f',   # 2025/01/01 00:14:00.000
            '%Y/%m/%d %H:%M:%S',      # 2025/01/01 00:14:00
        ]
        
        for time_str in time_column:
            if pd.isna(time_str):
                continue
            
            time_str = str(time_str).strip()
            if not time_str:
                continue
            
            # 尝试各种格式
            parsed = None
            for fmt in time_formats:
                try:
                    parsed = datetime.strptime(time_str, fmt)
                    break
                except:
                    continue
            
            # 如果格式解析失败，尝试pandas的自动解析
            if parsed is None:
                try:
                    parsed = pd.to_datetime(time_str, errors='coerce')
                    if pd.isna(parsed):
                        continue
                    parsed = parsed.to_pydatetime()
                except:
                    continue
            
            # 检查日期是否在指定范围内
            date_only = parsed.date()
            if self.start_date <= date_only <= self.end_date:
                timestamps.append(parsed)
        
        # 排序
        timestamps.sort()
        
        return timestamps
    
    def _calculate_daily_stats(self, timestamps: List[datetime]) -> Dict:
        """
        按天统计数据量
        
        Args:
            timestamps: 时间戳列表
            
        Returns:
            按天统计的字典，键为日期字符串，值为该天的记录数
        """
        daily_stats = {}
        
        for ts in timestamps:
            date_str = ts.date().strftime('%Y-%m-%d')
            daily_stats[date_str] = daily_stats.get(date_str, 0) + 1
        
        return daily_stats
    
    def _find_missing_dates(self, daily_stats: Dict) -> List[str]:
        """
        查找缺失的日期
        
        Args:
            daily_stats: 按天统计的字典
            
        Returns:
            缺失的日期列表（字符串格式：YYYY-MM-DD）
        """
        missing_dates = []
        current_date = self.start_date
        
        while current_date <= self.end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            if date_str not in daily_stats:
                missing_dates.append(date_str)
            current_date += timedelta(days=1)
        
        return missing_dates
    
    def _calculate_data_stats(self, data: pd.DataFrame) -> Dict:
        """
        计算数据统计信息
        
        Args:
            data: DataFrame
            
        Returns:
            统计信息字典
        """
        stats = {
            'total_rows': len(data),
            'total_columns': len(data.columns),
        }
        
        # 统计每列的数据
        column_stats = {}
        for i, col in enumerate(data.columns):
            col_data = data[col]
            col_stats = {
                'non_null_count': col_data.notna().sum(),
                'null_count': col_data.isna().sum(),
            }
            
            # 如果是数值列，计算统计信息
            if pd.api.types.is_numeric_dtype(col_data):
                col_stats.update({
                    'min': float(col_data.min()) if col_data.notna().any() else None,
                    'max': float(col_data.max()) if col_data.notna().any() else None,
                    'mean': float(col_data.mean()) if col_data.notna().any() else None,
                    'std': float(col_data.std()) if col_data.notna().any() else None,
                })
            
            column_stats[f'Column_{i}'] = col_stats
        
        stats['columns'] = column_stats
        
        return stats
    
    def generate_report(self, results: Dict, output_file: Optional[str] = None) -> str:
        """
        生成Excel检查报告
        
        Args:
            results: 检查结果字典
            output_file: 输出文件路径，如果为None则自动生成
            
        Returns:
            输出文件路径
        """
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = self.folder_path / f"时间完整性检查报告_{timestamp}.xlsx"
        else:
            output_file = Path(output_file)
        
        print(f"\n📝 正在生成检查报告: {output_file.name}")
        
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 创建样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. 创建汇总表
        self._create_summary_sheet(wb, results, header_fill, header_font, border)
        
        # 2. 为每个文件创建详细表
        for filename, result in results.items():
            if 'error' in result:
                continue
            self._create_file_detail_sheet(wb, filename, result, header_fill, header_font, border)
        
        # 3. 创建缺失时间汇总表
        self._create_missing_summary_sheet(wb, results, header_fill, header_font, border)
        
        # 保存文件
        wb.save(output_file)
        print(f"✅ 报告已保存: {output_file}")
        
        return str(output_file)
    
    def _create_summary_sheet(self, wb, results: Dict, header_fill, header_font, border):
        """创建汇总表"""
        ws = wb.create_sheet("文件汇总", 0)
        
        # 表头
        headers = [
            "文件名", "总记录数", "缺失天数", "缺失率(%)", 
            "最早日期", "最晚日期", "有数据天数", 
            "平均每天记录数", "最大每天记录数", "最小每天记录数", "状态"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 数据行
        row = 2
        for filename, result in results.items():
            if 'error' in result:
                ws.cell(row=row, column=1, value=filename).border = border
                ws.cell(row=row, column=11, value=f"错误: {result['error']}").border = border
                row += 1
                continue
            
            total = result['total_records']
            missing = result['missing_count']
            daily_stats = result.get('daily_stats', {})
            
            # 计算缺失率（基于天数）
            expected_days = (self.end_date - self.start_date).days + 1
            actual_days = len(daily_stats)
            missing_rate = (missing / expected_days * 100) if expected_days > 0 else 0
            
            # 计算日期范围
            dates_list = sorted(daily_stats.keys())
            first_date = dates_list[0] if dates_list else ''
            last_date = dates_list[-1] if dates_list else ''
            days_with_data = len(daily_stats)
            
            # 计算每天记录数统计
            daily_counts = list(daily_stats.values())
            avg_daily = np.mean(daily_counts) if daily_counts else 0
            max_daily = max(daily_counts) if daily_counts else 0
            min_daily = min(daily_counts) if daily_counts else 0
            
            data = [
                filename,
                total,
                missing,
                f"{missing_rate:.2f}",
                first_date,
                last_date,
                days_with_data,
                f"{avg_daily:.1f}",
                max_daily,
                min_daily,
                "正常" if missing == 0 else "有缺失"
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                if col_idx in [2, 3, 4, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 如果有缺失，标记行颜色
            if missing > 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _create_file_detail_sheet(self, wb, filename: str, result: Dict, 
                                  header_fill, header_font, border):
        """创建文件详细表"""
        # 清理工作表名称（Excel限制31个字符）
        sheet_name = filename[:31] if len(filename) > 31 else filename
        ws = wb.create_sheet(sheet_name)
        
        # 基本信息
        ws.cell(row=1, column=1, value="文件名:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=filename)
        ws.cell(row=2, column=1, value="总记录数:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=result['total_records'])
        ws.cell(row=3, column=1, value="缺失天数:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=result['missing_count'])
        
        # 按天统计
        daily_stats = result.get('daily_stats', {})
        row = 5
        ws.cell(row=row, column=1, value="按天统计信息").font = Font(bold=True, size=12)
        row += 1
        
        dates_list = sorted(daily_stats.keys())
        if dates_list:
            first_date = dates_list[0]
            last_date = dates_list[-1]
            days_with_data = len(daily_stats)
            daily_counts = list(daily_stats.values())
            avg_daily = np.mean(daily_counts) if daily_counts else 0
            max_daily = max(daily_counts) if daily_counts else 0
            min_daily = min(daily_counts) if daily_counts else 0
            
            time_info = [
                ("最早日期", first_date),
                ("最晚日期", last_date),
                ("有数据天数", days_with_data),
                ("平均每天记录数", f"{avg_daily:.1f}"),
                ("最大每天记录数", max_daily),
                ("最小每天记录数", min_daily),
            ]
            
            for label, value in time_info:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        # 数据统计
        data_stats = result.get('data_stats', {})
        row += 1
        ws.cell(row=row, column=1, value="数据统计信息").font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="总行数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data_stats.get('total_rows', 0))
        row += 1
        ws.cell(row=row, column=1, value="总列数").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data_stats.get('total_columns', 0))
        row += 1
        
        # 缺失日期详情
        missing_dates = result.get('missing_dates', [])
        if missing_dates:
            row += 1
            ws.cell(row=row, column=1, value="缺失日期详情").font = Font(bold=True, size=12)
            row += 1
            
            # 表头
            headers = ["序号", "缺失日期"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row += 1
            
            # 数据
            for idx, date_str in enumerate(missing_dates, 1):
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=date_str).border = border
                row += 1
        
        # 每天数据量详情
        if daily_stats:
            row += 1
            ws.cell(row=row, column=1, value="每天数据量详情").font = Font(bold=True, size=12)
            row += 1
            
            # 表头
            headers = ["日期", "记录数"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            row += 1
            
            # 数据（按日期排序）
            for date_str in sorted(daily_stats.keys()):
                ws.cell(row=row, column=1, value=date_str).border = border
                ws.cell(row=row, column=2, value=daily_stats[date_str]).border = border
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_missing_summary_sheet(self, wb, results: Dict, header_fill, header_font, border):
        """创建缺失日期汇总表"""
        ws = wb.create_sheet("缺失日期汇总")
        
        # 收集所有缺失日期
        all_missing = []
        for filename, result in results.items():
            if 'error' in result:
                continue
            for date_str in result.get('missing_dates', []):
                all_missing.append({
                    'filename': filename,
                    'date': date_str
                })
        
        if not all_missing:
            ws.cell(row=1, column=1, value="✅ 所有文件时间完整，无缺失日期")
            return
        
        # 按日期排序
        all_missing.sort(key=lambda x: x['date'])
        
        # 表头
        headers = ["序号", "文件名", "缺失日期"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 数据行
        for row_idx, item in enumerate(all_missing, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=item['filename']).border = border
            ws.cell(row=row_idx, column=3, value=item['date']).border = border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        # 冻结首行
        ws.freeze_panes = 'A2'


def main():
    """测试函数"""
    import sys
    
    if len(sys.argv) < 4:
        print("用法: python time_completeness_checker.py <文件夹路径> <开始日期> <结束日期>")
        print("示例: python time_completeness_checker.py ./data 2025-01-01 2025-01-31")
        return
    
    folder_path = sys.argv[1]
    start_date = sys.argv[2]
    end_date = sys.argv[3]
    
    try:
        checker = TimeCompletenessChecker(folder_path, start_date, end_date)
        results = checker.check_all_files()
        report_path = checker.generate_report(results)
        print(f"\n🎉 检查完成！报告已保存: {report_path}")
    except Exception as e:
        print(f"❌ 检查失败: {str(e)}")


if __name__ == "__main__":
    main()

